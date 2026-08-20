#!/usr/bin/env python3
"""Genera src/test/fixtures/sample.xlsx: un libro con el mismo formato que el
Excel original pero con datos inventados, para poder probar el importador en CI.

    pip install openpyxl && python3 src/test/fixtures/make-sample.py
"""
import os
import openpyxl

# columna de la cabecera de cada categoria (la de al lado lleva el importe
# y la siguiente el =SUM(), igual que en el Excel original)
LAYOUT = [
    ("外食", 2),                    # B/C, suma en D
    ("スーパーマーケット", 5),        # E/F, suma en G
    ("服装と電車と毎月費消", 8),      # H/I, suma en J
    ("娯楽", 11),                   # K/L, suma en M
    ("部屋のもの", 14),              # N/O, suma en P
]

def col(n: int) -> str:
    return openpyxl.utils.get_column_letter(n)

def write_headers(ws):
    ws["A1"] = "合計"
    for name, c in LAYOUT:
        ws.cell(row=1, column=c, value=name)
        ws.cell(row=1, column=c + 2, value=f"=SUM({col(c + 1)}:{col(c + 1)})")
    ws.cell(row=1, column=15, value="値段")  # etiqueta de la columna de precios

def write_items(ws, items):
    """items: {columna de cabecera: [(concepto, importe), ...]}"""
    for c, rows in items.items():
        for i, (label, amount) in enumerate(rows, start=2):
            ws.cell(row=i, column=c, value=label)
            ws.cell(row=i, column=c + 1, value=amount)

wb = openpyxl.Workbook()

# --- hoja 1: marzo de 2026, completa -------------------------------------
ws = wb.active
ws.title = "26-3月の費消"
write_headers(ws)
write_items(
    ws,
    {
        2: [("mcdonals", 1000), ("uber", 2000)],
        5: [("seiyu", 3000), ("lawson", 500)],
        8: [("suica", 10000), ("netflix", 1590)],
        11: [("steam", 2080)],
        14: [("ikea", 22990)],
    },
)
ws["A2"] = "=ROUNDUP(M1+J1+G1+D1+P1+A5+A12,0)"
ws["A4"] = "家賃"
ws["A5"] = 82000
ws["A7"] = "一日生活の費消/支出"
ws["A10"] = "毎月ある費消/衣食住"
ws["A12"] = "=1000+2000"
ws["A14"] = "為替相場"
ws["A15"] = 0.0056
ws["A19"] = "上限・バランス"
ws["A20"] = 200000
ws["A22"] = "nota de prueba en la hoja"
ws["Q1"] = "hotel"
ws["R1"] = 5000

# --- hoja 2: diciembre, con el limite escrito en euros -------------------
ws2 = wb.create_sheet("十二月")
write_headers(ws2)
write_items(ws2, {2: [("kokosu", 990)], 5: [("famima", 670)]})
ws2["A4"] = "家賃"
ws2["A5"] = 91000
ws2["A14"] = "為替相場"
ws2["A15"] = 0.0056
ws2["A19"] = "上限・バランス"
ws2["A20"] = 1300  # euros: el importador lo pasa a yenes

# --- hoja 3: plantilla sin datos (debe descartarse) ---------------------
ws3 = wb.create_sheet("plantilla vacia")
write_headers(ws3)

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sample.xlsx")
wb.save(out)
print("escrito", out)
